from openpyxl import load_workbook
from datetime import datetime, timedelta
from manufacture.models import Production
from manufacture.services.write_to_report import write_to_report
from itertools import groupby
from operator import itemgetter

def report_production(attachment):
    current_month_year = datetime.now().strftime('%m') + '.' + datetime.now().strftime('%Y')
    # current_month_year = '07.2024' # Для тестов

    yesterday_date = datetime.now() - timedelta(days=1)
    yesterday_date = yesterday_date.replace(hour=0, minute=0, second=0, microsecond=0)
    # yesterday_date = '2024-04-16 00:00:00'
    # yesterday_date = datetime.strptime(yesterday_date, "%Y-%m-%d %H:%M:%S")
    
    file_path = attachment.document
    print(file_path)

    workbook = load_workbook(filename=file_path)
    product_titles = []
    product_values = []
    
    digital_product_titles = []
    digital_product_values = []

    digital_product_varients = []
    digital_product_weights = []

    text_data = ''

    for sheet in workbook.worksheets:
        if current_month_year in sheet.title:
            print(sheet)
            date_range = sheet['A2':'A32']
            product_range = sheet['C1':'M1']

            for data_cell in date_range:
                if data_cell[0].value.replace(year=datetime.now().year) == yesterday_date:
                    for product in product_range[0]:
                        product_titles.append(product.value)

                        product_value = sheet.cell(row=data_cell[0].row, column=product.column).value
                        product_values.append(product_value)
                        

    excel_data = [{'name':a, 'weight':b} for a,b in zip(product_titles, product_values)]
    excel_data = [elem for elem in excel_data if elem['weight'] != None]


    digital_data = Production.objects.filter(created_at__year = yesterday_date.year,
                                             created_at__month = yesterday_date.month,
                                             created_at__day = yesterday_date.day)


    for item in digital_data:
        if item.product not in digital_product_varients:
            digital_product_varients.append(item.product)

    
    for item in digital_product_varients:
        digital_product_titles.append(item.name)
        item_objects = digital_data.filter(product = item)

        sum_weights = 0

        for itemm in item_objects: 
            sum_weights += int(itemm.weight)


        digital_product_weights.append(sum_weights)

    digital_data = [{'name':a, 'weight':b} for a,b in zip(digital_product_titles, digital_product_weights)]

    for item in excel_data:
        match item['name']:
            case 'Плёнка микс':
                item['name'] = 'Пленка ПВД неокрашенная'
                
            case 'Пленка ПВД':
                item['name'] = 'Пленка'
            case 'бумага':
                item['name'] = 'Бумага'
            case 'ПЭТ ПГ':
                item['name'] = 'ПЭТ ПГ'
            case 'ПЭТ микс':
                item['name'] = 'ПЭТ микс'
                april_value = 3569
            case 'ПНД':
                item['name'] = 'ПНД'
                april_value = 76
            case 'ПЭТ б.':
                item['name'] = 'ПЭТ б.'
            case 'Тетра Пак':
                item['name'] = 'Тетра Пак'
            case 'Алюм.':
                item['name'] = 'Алюминий'
                april_value = 304
            case 'Стекло':
                item['name'] = 'Стекло'
                april_value = 470
            case 'Железо':
                item['name'] = 'Черный металл'
                april_value = 1355

    for digital_item in digital_data:
        match digital_item['name']:
            case 'ПЭТ микс':
                april_value = 3569
            case 'ПНД':
                april_value = 76
            case 'Алюминий':
                april_value = 304
            case 'Стекло':
                april_value = 1356
            case 'Черный металл':
                april_value = 1355
            case _:
                april_value = 0

        for excel_item in excel_data:
            if excel_item['name'] == digital_item['name']:
                text_data += digital_item['name'] + ': ' + str(excel_item['weight']) + ' / ' + str(digital_item['weight']) + ' / ' + str(april_value) +'\n'

    # print('digital_data: ', digital_data)
    # print('excel_data: ', excel_data)
    # print('text_data: ', text_data)

    write_to_report(7, yesterday_date, text_data)