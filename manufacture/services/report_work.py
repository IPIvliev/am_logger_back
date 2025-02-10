from openpyxl import load_workbook
from datetime import datetime, timedelta
from manufacture.models import StopReport
from django.db.models import Sum
from django.db.models import F, IntegerField, ExpressionWrapper, CharField
from django.db.models.functions import Cast, Extract
from manufacture.services.write_to_report import write_to_report

def report_work(attachment, total_minutes):
    file_path = attachment.document
    workbook = load_workbook(filename=file_path, data_only=True)

    # windows
    # yesterday_day_month_without_zero = (datetime.now() - timedelta(days=1)).strftime('%#d') + '.' + (datetime.now()- timedelta(days=1)).strftime('%m')

    # linux
    # yesterday_day_month_without_zero = (datetime.now() - timedelta(days=1)).strftime('%-d') + '.' + (datetime.now()- timedelta(days=1)).strftime('%m')

    yesterday_day_month_with_zero = (datetime.now() - timedelta(days=1)).strftime('%d') + '.' + (datetime.now()- timedelta(days=1)).strftime('%m')
    yesterday_date = datetime.now() - timedelta(days=1)
    yesterday_date = yesterday_date.replace(hour=0, minute=0, second=0, microsecond=0)

    excel_reasons_night = []
    excel_values_night = []
    excel_reasons_day = []
    excel_values_day = []

    def search_for_data_range(sheet, column):
        # print('this row 76: ', (sheet.cell(row=76, column=column).value))
        # print('this row 77: ', (sheet.cell(row=77, column=column).value))
        # print('this row 78: ', (sheet.cell(row=78, column=column).value))
        # print('this row 79: ', (sheet.cell(row=79, column=column).value))

        if (type(sheet.cell(row=76, column=column).value) is str):
            # print('this row 76: ', (sheet.cell(row=76, column=column).value))
            sheet.cell(row=76, column=column).value = 0
            # print('this row 76: ', (sheet.cell(row=76, column=column).value))
        if (type(sheet.cell(row=77, column=column).value) is str):
            # print('this row 77: ', (sheet.cell(row=77, column=column).value))
            sheet.cell(row=77, column=column).value = 0
            # print('this row 77: ', (sheet.cell(row=77, column=column).value))
        if (type(sheet.cell(row=78, column=column).value) is str):
            # print('this row 78: ', (sheet.cell(row=78, column=column).value))
            sheet.cell(row=78, column=column).value = 0
            # print('this row 78: ', (sheet.cell(row=78, column=column).value))
        if (type(sheet.cell(row=79, column=column).value) is str):
            # print('this row 79: ', (sheet.cell(row=79, column=column).value))
            sheet.cell(row=79, column=column).value = 0
            # print('this row 79: ', (sheet.cell(row=79, column=column).value))

        # try:
        if (sheet.cell(row=76, column=column).value != None) and (sheet.cell(row=76, column=column).value > 30):
            print(sheet.cell(row=76, column=column).value)
            if column == 3:
                data_range = sheet['C76':'O76']
            elif column == 2:
                data_range = sheet['B76':'B76']
            return data_range
        elif (sheet.cell(row=77, column=column).value != None) and (sheet.cell(row=77, column=column).value > 30):
            print(sheet.cell(row=77, column=column).value)
            if column == 3:
                data_range = sheet['C77':'O77']
            elif column == 2:
                data_range = sheet['B77':'B77']
            return data_range
        elif (sheet.cell(row=78, column=column).value != None) and (sheet.cell(row=78, column=column).value > 30):
            print(sheet.cell(row=78, column=column).value)
            if column == 3:
                data_range = sheet['C78':'O78']
            elif column == 2:
                data_range = sheet['B78':'B78']
            return data_range
        elif (sheet.cell(row=79, column=column).value != None) and (sheet.cell(row=79, column=column).value > 30):
            print(sheet.cell(row=79, column=column).value)
            if column == 3:
                data_range = sheet['C79':'O79']
            elif column == 2:
                data_range = sheet['B79':'B79']
            return data_range
        else:
            # print(0)
            data_range = sheet['C179':'O179']
            return data_range
        # except:
        #     data_range = sheet['C179':'O179']
        #     return data_range

    def search_for_reason_range(sheet):
        # print('3 ', sheet.cell(row=3, column=4).value)
        # print('4 ', sheet.cell(row=4, column=2).value)
        if "Работа" in str(sheet.cell(row=3, column=3).value):
            data_range = search_for_data_range(sheet, 3)
            reason_range = sheet['C3':'O3']
            # print(sheet.cell(row=3, column=2).value)
            return reason_range, data_range
        elif "Работа" in str(sheet.cell(row=4, column=3).value):
            data_range = search_for_data_range(sheet, 3)
            reason_range = sheet['C4':'O4']
            # print(sheet.cell(row=3, column=2).value)
            return reason_range, data_range
        elif "Работа" in str(sheet.cell(row=3, column=2).value):
            data_range = search_for_data_range(sheet, 2)
            reason_range = sheet['B3':'O3']
            # print(sheet.cell(row=3, column=2).value)
            return reason_range, data_range
        elif "Работа" in str(sheet.cell(row=4, column=2).value):
            data_range = search_for_data_range(sheet, 2)
            reason_range = sheet['B4':'O4']
            # print(sheet.cell(row=3, column=2).value)
            return reason_range, data_range
        elif "Работа" in str(sheet.cell(row=3, column=4).value):
            data_range = search_for_data_range(sheet, 2)
            reason_range = sheet['D3':'P3']
            # print(sheet.cell(row=3, column=2).value)
            return reason_range, data_range
          
    for sheet in workbook.worksheets:
        # print(sheet, yesterday_day_month_with_zero, yesterday_day_month_without_zero)
        title = sheet.title
        title = title.split('.')

        if len(title[0]) == 1:
            title[0] = '0' + title[0]

        if len(title) == 2:
            title = title[0] + '.' + title[1].lower()
        elif len(title) == 3:
            title = title[0] + '.' + title[1].lower() + title[2].lower()


        if yesterday_day_month_with_zero in title:

            if 'ночь' in title:
                print('ночь: ', title)
                reason_range, data_range = search_for_reason_range(sheet)
                for data_cell in reason_range[0]:
                    excel_reasons_night.append(data_cell.value)
                for data_cell in data_range[0]:
                    excel_values_night.append(data_cell.value)         

            elif 'день' in title:
                print('день: ', title)
                reason_range, data_range = search_for_reason_range(sheet)

                for data_cell in reason_range[0]:
                    excel_reasons_day.append(data_cell.value)
                for data_cell in data_range[0]:
                    excel_values_day.append(data_cell.value)  

    digital_data = StopReport.objects.filter(created_at__year = yesterday_date.year,
                                        created_at__month = yesterday_date.month,
                                        created_at__day = yesterday_date.day)

    # stop = digital_data.annotate(stop_period=ExpressionWrapper(Cast(F('created_at') - F('finished_at'), CharField()) , output_field=CharField()))
    stop = digital_data.annotate(stop_period=(F('finished_at') - F('created_at'))).aggregate(Sum('stop_period'))

    try:
        digital_hours, digital_minutes = str(timedelta(minutes=total_minutes) - stop['stop_period__sum']).split(':', 2)[:2]
    except:
        digital_hours, digital_minutes = str(timedelta(minutes=(total_minutes - 240))).split(':', 2)[:2]

    digital_data = str(digital_hours + ':' + digital_minutes)

    # print('excel_values_night[0] above checks', excel_values_night[0])

    if excel_values_day[0] == None:
        # print('excel_values_day[0] ', excel_values_day[0])
        excel_values_day[0] = 0

    if (excel_values_night[0] == None):
        # print('excel_valuessd sadas dd_night[0] ', excel_values_night[0])
        excel_values_night[0] = 0

    print('excel_values_day[0] ', excel_values_day[0])
    print('excel_values_night[0] ', excel_values_night[0])

    excel_hours, excel_minutes = str(timedelta(minutes=(excel_values_day[0] + excel_values_night[0]))).split(':', 2)[:2]
    excel_data = str(excel_hours + ':' + excel_minutes)

    text_data = excel_data + ' / ' + digital_data + ' / ' + '20:00'

    # print(text_data)

    write_to_report(2, yesterday_date, excel_data)
    write_to_report(3, yesterday_date, digital_data)
    write_to_report(4, yesterday_date, '20:00')