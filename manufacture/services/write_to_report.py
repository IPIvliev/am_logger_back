from openpyxl import load_workbook
import os
from datetime import datetime, timedelta

directory = 'E:/sites/am_logger_back/public_html/media/mail_reports/'
# directory = '/home/razrus/am_logger_back/public_html/media/mail_reports/'

current_year = str(datetime.now().year)

def write_to_report(column, yesterday_date, data):
    
    # yesterday_month_year = datetime.now().strftime('%m') + '.' + datetime.now().strftime('%Y')
    yesterday_month_year = (datetime.now() - timedelta(days=1)).strftime('%m') + '.' + (datetime.now() - timedelta(days=1)).strftime('%Y')

    for root, directories, files in os.walk(directory):
        for filename in files:
            if current_year in filename:
                file_path = os.path.join(root, filename)
                # print(file_path)

    workbook = load_workbook(filename=file_path)

    for sheet in workbook.worksheets:
        # print(yesterday_month_year)
        

        if yesterday_month_year in sheet.title:
            date_range = sheet['A6':'A36']
            print(sheet)
            for data_cell in date_range:
                # print("only_date", data_cell[0].value.date())
                try:
                    only_date, only_time = data_cell[0].value.date(), data_cell[0].value.time()
                    
                    data_cell[0].value = only_date
                    # cell = sheet.cell(row=data_cell[0].row, column=data_cell[0].column)
                    # cell.number_format = 'DD.MM.YYYY'

                    if data_cell[0].value == yesterday_date.date():
                        row = data_cell[0].row

                        # sheet.cell(row=row, column=column).value = str(data).replace(',', '\n').replace('{', '').replace('}', '')
                        sheet.cell(row=row, column=column).value = data
                        # print("get")
                        
                        workbook.save(file_path)
                    
                
                    # data_cell[0].value = datetime.strptime(data_cell[0].value, "%Y.%m.%d")
                except:
                    print('Smth wrong')


