from datetime import datetime, timedelta
from manufacture.models import Parameter
from manufacture.services.write_to_report import write_to_report
from django.db.models import Avg
from django.db.models.functions import Cast
from django.db.models import IntegerField
from openpyxl import load_workbook
import locale

def add_personal(attachment):
    locale.setlocale(locale.LC_TIME, 'ru_RU')

    yesterday_date = datetime.now() - timedelta(days=1)
    yesterday_date = yesterday_date.replace(hour=0, minute=0, second=0, microsecond=0)


    yesterday_month_year = (datetime.now() - timedelta(days=1)).strftime('%B') + ' ' + (datetime.now() - timedelta(days=1)).strftime('%Y')
    # month = datetime.strptime(str(yesterday_date), "%d.%m.%Y").strftime("%B")

    # if 'января' in yesterday_month_year:
    yesterday_month_year = yesterday_month_year.replace("января", "январь")
    yesterday_month_year = yesterday_month_year.replace("февраля", "февраль")
    yesterday_month_year = yesterday_month_year.replace("марта", "март")
    yesterday_month_year = yesterday_month_year.replace("апреля", "апрель")
    yesterday_month_year = yesterday_month_year.replace("мая", "май")
    yesterday_month_year = yesterday_month_year.replace("июня", "июнь")
    yesterday_month_year = yesterday_month_year.replace("июля", "июль")
    yesterday_month_year = yesterday_month_year.replace("августа", "август")
    yesterday_month_year = yesterday_month_year.replace("сентября", "сентябрь")
    yesterday_month_year = yesterday_month_year.replace("октября", "октябрь")
    yesterday_month_year = yesterday_month_year.replace("ноября", "ноябрь")
    yesterday_month_year = yesterday_month_year.replace("декабря", "декабрь")

    print(yesterday_month_year)
    # print(month)

    file_path = attachment.document
    workbook = load_workbook(filename=file_path, data_only=True)

    for sheet in workbook.worksheets:
        if yesterday_month_year.lower() in sheet.title:
            # print(yesterday_month_year)
            date_range = sheet['A3':'A32']

            for data_cell in date_range:
                try:
                    if data_cell[0].value.replace(year=datetime.now().year) == yesterday_date:
                        personal_fact_value = sheet.cell(row=data_cell[0].row, column=2).value
                        # print(personal_fact_value)
                        personal_norm_value = sheet.cell(row=data_cell[0].row, column=3).value
                        # print(personal_norm_value)
                except:
                    personal_fact_value = 21
                    personal_norm_value = 29

    write_to_report(5, yesterday_date, personal_fact_value)
    write_to_report(6, yesterday_date, personal_norm_value)